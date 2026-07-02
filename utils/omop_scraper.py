#!/usr/bin/env python3
"""
omop_scraper.py  –  OMOP CDM full-hierarchy builder
====================================================
Produces:  TableGroup → Table → Column
           (parallel to PhenX: Collection → SubCollection → Protocol → Variable)

Data source: two canonical CSV files from the OHDSI CommonDataModel GitHub repo
  OMOP_CDMv{version}_Table_Level.csv  — one row per table
  OMOP_CDMv{version}_Field_Level.csv  — one row per column

Table groups are discovered at runtime by fetching the cdm{version}.Rmd source
from GitHub, which contains the authoritative group-boundary trigger tables
(the same logic used to render the official OHDSI docs page).
If that fetch fails, the CSV `schema` field is used as a coarse fallback.

QUICK START
-----------
    pip install requests openpyxl
    python omop_scraper.py                        # CDM v5.4 → omop_5.4.xlsx
    python omop_scraper.py --version 5.4          # explicit → omop_5.4.xlsx
    python omop_scraper.py --version 5.4 --out data/omop_5.4.xlsx

OUTPUT
------
    omop_{version}.xlsx  – README | Tables | Columns | Reference
    omop_{version}.csv   – flat column rows (same data as Columns sheet)
"""

import argparse
import csv
import io
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
GITHUB_RAW      = "https://raw.githubusercontent.com/OHDSI/CommonDataModel"
DOCS_BASE       = "https://ohdsi.github.io/CommonDataModel"
BROWSER_UA      = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
DEFAULT_VERSION = "5.4"


# ── HTTP ──────────────────────────────────────────────────────────────────────

def _session():
    s = requests.Session()
    s.headers.update({"User-Agent": BROWSER_UA})
    return s


def _get_text(session, url, retries=3, delay=1.0):
    for attempt in range(1, retries + 1):
        time.sleep(delay)
        try:
            r = session.get(url, timeout=30)
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", 60))
                print(f"\n  [rate-limit] sleeping {wait}s", file=sys.stderr)
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.text
        except requests.RequestException as e:
            if attempt == retries:
                print(f"\n  [error] {url}: {e}", file=sys.stderr)
                return None
            time.sleep(delay * 3)
    return None


# ── Phase 0: fetch canonical CSVs from OHDSI GitHub ──────────────────────────

def fetch_cdm_data(session, version, delay):
    """
    Fetch Table_Level and Field_Level CSVs from the OHDSI CommonDataModel repo.
    Tag format: v{version}.0  (e.g. v5.4.0).
    Returns (table_rows, field_rows) — lists of dicts, in CSV row order.
    """
    tag      = f"v{version}.0"
    base_url = f"{GITHUB_RAW}/{tag}/inst/csv"

    print(f"Fetching Table_Level CSV (v{version}) …", end=" ", flush=True)
    table_text = _get_text(session, f"{base_url}/OMOP_CDMv{version}_Table_Level.csv", delay=delay)
    if not table_text:
        raise RuntimeError("Could not fetch Table_Level CSV")
    tables = list(csv.DictReader(io.StringIO(table_text)))
    print(f"{len(tables)} tables")

    print(f"Fetching Field_Level CSV (v{version}) …", end=" ", flush=True)
    field_text = _get_text(session, f"{base_url}/OMOP_CDMv{version}_Field_Level.csv", delay=delay)
    if not field_text:
        raise RuntimeError("Could not fetch Field_Level CSV")
    fields = list(csv.DictReader(io.StringIO(field_text)))
    print(f"{len(fields)} fields")

    return tables, fields


# ── Phase 1: resolve table groups from Rmd source ────────────────────────────

def fetch_group_boundaries(session, version, delay):
    """
    Fetch cdm{version}.Rmd from GitHub and extract the group-boundary trigger
    tables — the same if/cat blocks used to render the official OHDSI docs.

    Returns dict: trigger_table_upper → group_name
    e.g. {'PERSON': 'Clinical_Data_Tables', 'LOCATION': 'Health_System_Data_Tables', ...}
    """
    tag     = f"v{version}.0"
    ver_str = version.replace(".", "")   # "5.4" → "54"
    rmd_url = f"{GITHUB_RAW}/{tag}/rmd/cdm{ver_str}.Rmd"

    print(f"Fetching cdm{ver_str}.Rmd for group taxonomy …", end=" ", flush=True)
    text = _get_text(session, rmd_url, delay=delay)
    if not text:
        print("FAILED — falling back to schema field")
        return {}

    # Pattern: if(tb == 'TABLE'){ cat("## **Group Name**\n\n") }
    # We look for  if(tb == 'X'){  followed soon by  cat("## **Y**
    pat = re.compile(
        r"if\s*\(\s*tb\s*==\s*'([A-Z_]+)'\s*\)\s*\{[^}]*"
        r'cat\s*\(\s*"##\s*\*\*([^*]+)\*\*',
        re.DOTALL
    )
    boundaries = {}
    for m in pat.finditer(text):
        table      = m.group(1).strip().upper()
        group_raw  = m.group(2).strip()
        # Normalise: "Clinical Data Tables" → "Clinical_Data_Tables"
        group_name = re.sub(r'\s+', '_', group_raw)
        boundaries[table] = group_name

    print(f"{len(boundaries)} boundaries found: {list(boundaries.values())}")
    return boundaries


def resolve_groups(tables, session, version, delay):
    """
    Assign each table a group name.

    Priority:
      1. Rmd group-boundary scan (authoritative, matches OHDSI docs page h2 headings)
      2. CSV `schema` field as coarse fallback (CDM/VOCAB/RESULTS → *_Tables)

    Tables whose schema=RESULTS (COHORT, COHORT_DEFINITION) are not in the Rmd
    loop so they get schema-field fallback → Results_Tables.

    Returns dict: table_name_upper → group_name
    """
    boundaries = fetch_group_boundaries(session, version, delay)

    # Walk tables in CSV order, carrying current group forward from each boundary
    current_group = None
    t2g = {}
    for row in tables:
        tname  = row["cdmTableName"].strip().upper()
        schema = row.get("schema", "CDM").strip().upper()

        if tname in boundaries:
            current_group = boundaries[tname]

        if current_group:
            t2g[tname] = current_group
        else:
            # Fallback: schema field
            t2g[tname] = f"{schema.capitalize()}_Tables"

        # Schema=RESULTS tables are not in the Rmd loop — override
        if schema == "RESULTS":
            t2g[tname] = "Results_Tables"

    return t2g


# ── Phase 2: build structured rows ───────────────────────────────────────────

def build_rows(tables, fields, t2g, version):
    """
    Build one row per table (rows_t) and one row per column (rows_c).
    All values come from the CSV — nothing is hardcoded.

    Column ID format:  OMOPColumn_{TABLE}__{field}   (parallel to PX variable IDs)
    """
    t_meta   = {r["cdmTableName"].strip().upper(): r for r in tables}
    doc_slug = f"cdm{version.replace('.', '')}"   # "5.4" → "cdm54"

    rows_t = []
    rows_c = []

    for tname, trow in t_meta.items():
        group    = t2g.get(tname, "CDM_Tables")
        schema   = trow.get("schema", "").strip()
        t_desc   = trow.get("tableDescription", "").strip()
        u_guide  = trow.get("userGuidance", "").strip()
        etl_conv = trow.get("etlConventions", "").strip()
        is_req   = trow.get("isRequired", "").strip()
        tbl_url  = f"{DOCS_BASE}/{doc_slug}.html#{tname.lower()}"

        tbl_cols = [f for f in fields if f["cdmTableName"].strip().upper() == tname]

        rows_t.append({
            "table_group":       group,
            "table_name":        tname,
            "schema":            schema,
            "is_required":       is_req,
            "table_description": t_desc[:600],
            "user_guidance":     u_guide[:400],
            "etl_conventions":   etl_conv[:400],
            "n_columns":         len(tbl_cols),
            "table_url":         tbl_url,
            "cdm_version":       version,
        })

        for col in tbl_cols:
            col_name = col.get("cdmFieldName",    "").strip()
            col_req  = col.get("isRequired",      "").strip()
            col_type = col.get("cdmDatatype",     "").strip()
            col_guid = col.get("userGuidance",    "").strip()
            col_etl  = col.get("etlConventions",  "").strip()
            is_pk    = col.get("isPrimaryKey",    "").strip()
            is_fk    = col.get("isForeignKey",    "").strip()
            fk_table = col.get("fkTableName",     "").strip()
            fk_field = col.get("fkFieldName",     "").strip()
            fk_dom   = col.get("fkDomain",        "").strip()
            fk_class = col.get("fkClass",         "").strip()

            # Unique column ID: OMOPColumn_{TABLE}__{field}
            col_id = f"OMOPColumn_{tname}__{col_name}"

            rows_c.append({
                "column_id":          col_id,
                "column_name":        col_name,
                "column_description": col_guid[:600] if col_guid else col_name,
                "etl_conventions":    col_etl[:400],
                "datatype":           col_type,
                "is_required":        col_req,
                "is_primary_key":     is_pk,
                "is_foreign_key":     is_fk,
                "fk_table":           fk_table,
                "fk_field":           fk_field,
                "fk_domain":          fk_dom,
                "fk_class":           fk_class,
                "table_name":         tname,
                "table_group":        group,
                "table_description":  t_desc[:300],
                "table_url":          tbl_url,
                "schema":             schema,
                "cdm_version":        version,
            })

    return rows_t, rows_c


# ── Excel builder ─────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
WHITE    = PatternFill("solid", start_color="FFFFFF")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BDY_FONT = Font(name="Arial", size=9)
BLD_FONT = Font(name="Arial", bold=True, size=9)
LNK_FONT = Font(name="Arial", size=9, color="1155CC", underline="single")
_thin    = Side(style="thin", color="BFBFBF")
BORDER   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP     = Alignment(vertical="top", wrap_text=True)

def _h(cell, text):
    cell.value     = text
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def _b(cell, val, alt=False, bold=False, link=False):
    cell.value     = val
    cell.font      = LNK_FONT if link else (BLD_FONT if bold else BDY_FONT)
    cell.fill      = ALT_FILL if alt else WHITE
    cell.alignment = WRAP
    cell.border    = BORDER

def _w(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def _title(ws, text, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c           = ws.cell(1, 1)
    c.value     = text
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def build_excel(rows_t, rows_c, version, out_path):
    wb     = Workbook()
    tag    = f"v{version}.0"
    n_grps = len(set(r["table_group"] for r in rows_t))

    # ── README ────────────────────────────────────────────────────────────────
    ws0       = wb.active
    ws0.title = "README"
    ws0.column_dimensions["A"].width = 26
    ws0.column_dimensions["B"].width = 90
    meta = [
        ("Source",          f"{DOCS_BASE}/cdm{version.replace('.','')}.html"),
        ("CDM Version",     f"OMOP CDM v{version}"),
        ("Generated",       time.strftime("%Y-%m-%d")),
        ("Table Groups",    str(n_grps)),
        ("Tables",          str(len(rows_t))),
        ("Columns / CDEs",  str(len(rows_c))),
        ("Hierarchy",       "TableGroup → Table → Column"),
        ("Sheets",          "README | Tables | Columns | Reference"),
        ("CDM R Package",   "https://github.com/OHDSI/CommonDataModel"),
        ("Table-level CSV", f"{GITHUB_RAW}/{tag}/inst/csv/OMOP_CDMv{version}_Table_Level.csv"),
        ("Field-level CSV", f"{GITHUB_RAW}/{tag}/inst/csv/OMOP_CDMv{version}_Field_Level.csv"),
        ("Group taxonomy",  f"{GITHUB_RAW}/{tag}/rmd/cdm{version.replace('.','')}.Rmd"),
        ("License",         "Apache 2.0 – OHDSI CommonDataModel"),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws0.cell(r, 1).value = k
        ws0.cell(r, 1).font  = Font(name="Arial", bold=True, size=10)
        ws0.cell(r, 2).value = v
        ws0.cell(r, 2).font  = Font(name="Arial", size=10)

    # ── Tables ────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Tables")
    ws1.freeze_panes = "A3"
    T_HDRS = [
        "Table Group", "Table Name", "Schema", "Is Required",
        "Table Description", "User Guidance", "ETL Conventions",
        "# Columns", "Table URL", "CDM Version",
    ]
    _title(ws1, f"OMOP CDM v{version} – Tables ({len(rows_t)})  |  {DOCS_BASE}", len(T_HDRS))
    ws1.row_dimensions[2].height = 18
    for c, h in enumerate(T_HDRS, 1):
        _h(ws1.cell(2, c), h)
    for i, w in enumerate([32, 28, 14, 12, 65, 55, 55, 10, 60, 12], 1):
        _w(ws1, i, w)
    for i, t in enumerate(rows_t):
        r = i + 3
        alt = i % 2 == 1
        vals = [
            t["table_group"],       t["table_name"],
            t["schema"],            t["is_required"],
            t["table_description"], t["user_guidance"],
            t["etl_conventions"],   str(t["n_columns"]),
            t["table_url"],         t["cdm_version"],
        ]
        for c, v in enumerate(vals, 1):
            _b(ws1.cell(r, c), v, alt, bold=(c == 2), link=(c == 9))
        ws1.row_dimensions[r].height = 36

    # ── Columns / CDEs ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Columns (CDEs)")
    ws2.freeze_panes = "A3"
    C_HDRS = [
        "Column ID (CDE ID)", "Column Name", "Column Description",
        "ETL Conventions", "Datatype", "Is Required",
        "Is Primary Key", "Is Foreign Key",
        "FK Table", "FK Field", "FK Domain", "FK Class",
        "Table Name", "Table Group", "Table Description", "Table URL",
        "Schema", "CDM Version",
    ]
    _title(ws2, f"OMOP CDM v{version} – Columns / CDEs ({len(rows_c)})  |  {DOCS_BASE}", len(C_HDRS))
    ws2.row_dimensions[2].height = 18
    for c, h in enumerate(C_HDRS, 1):
        _h(ws2.cell(2, c), h)
    for i, w in enumerate(
        [40, 32, 65, 55, 12, 12, 14, 14, 24, 20, 18, 16, 28, 32, 55, 60, 12, 12], 1
    ):
        _w(ws2, i, w)
    for i, col in enumerate(rows_c):
        r = i + 3
        alt = i % 2 == 1
        vals = [
            col["column_id"],          col["column_name"],
            col["column_description"], col["etl_conventions"],
            col["datatype"],           col["is_required"],
            col["is_primary_key"],     col["is_foreign_key"],
            col["fk_table"],           col["fk_field"],
            col["fk_domain"],          col["fk_class"],
            col["table_name"],         col["table_group"],
            col["table_description"],  col["table_url"],
            col["schema"],             col["cdm_version"],
        ]
        for c, v in enumerate(vals, 1):
            _b(ws2.cell(r, c), v, alt, bold=(c == 1), link=(c == 16))
        ws2.row_dimensions[r].height = 30

    # ── Reference ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Reference")
    ws3.freeze_panes = "A2"
    for c, h in enumerate(
        ["Table Group", "Schema", "Table Name", "Is Required", "# Columns", "Table URL"], 1
    ):
        _h(ws3.cell(1, c), h)
    ws3.row_dimensions[1].height = 18
    for i, w in enumerate([32, 12, 28, 12, 10, 65], 1):
        _w(ws3, i, w)
    for i, t in enumerate(rows_t):
        r = i + 2
        alt = i % 2 == 1
        _b(ws3.cell(r, 1), t["table_group"],   alt, bold=True)
        _b(ws3.cell(r, 2), t["schema"],         alt)
        _b(ws3.cell(r, 3), t["table_name"],     alt, bold=True)
        _b(ws3.cell(r, 4), t["is_required"],    alt)
        _b(ws3.cell(r, 5), str(t["n_columns"]), alt)
        _b(ws3.cell(r, 6), t["table_url"],      alt, link=True)
        ws3.row_dimensions[r].height = 20

    wb.save(out_path)
    print(f"Saved Excel → {out_path}")


def write_csv(rows_c, path):
    if not rows_c:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_c[0].keys()))
        w.writeheader()
        w.writerows(rows_c)
    print(f"Saved CSV   → {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Build OMOP CDM hierarchy → TableGroup → Table → Column",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--version", default=DEFAULT_VERSION,
                    help=f"CDM version to fetch (default: {DEFAULT_VERSION})")
    ap.add_argument("--delay",   type=float, default=0.5,
                    help="Seconds between HTTP requests (default 0.5)")
    ap.add_argument("--out",     default=None,
                    help="Output Excel path (default: omop_{version}.xlsx)")
    args = ap.parse_args()

    out_xlsx = Path(args.out) if args.out else Path(f"omop_{args.version}.xlsx")
    out_csv  = out_xlsx.with_suffix(".csv")
    session  = _session()

    print("=" * 60)
    print(f"OMOP CDM v{args.version} hierarchy builder")
    print("=" * 60)

    print("\nPhase 0: fetching canonical CDM CSVs from OHDSI GitHub …")
    tables, fields = fetch_cdm_data(session, args.version, args.delay)

    print("\nPhase 1: resolving table groups …")
    t2g = resolve_groups(tables, session, args.version, args.delay)
    by_group = defaultdict(list)
    for t, g in t2g.items():
        by_group[g].append(t)
    for grp in sorted(by_group):
        print(f"  {grp:<40}  {len(by_group[grp])} tables: {by_group[grp]}")

    print("\nPhase 2: building structured rows …")
    rows_t, rows_c = build_rows(tables, fields, t2g, args.version)
    print(f"  → {len(rows_t)} tables, {len(rows_c)} columns")

    print(f"\n{'='*60}\nWriting outputs …")
    build_excel(rows_t, rows_c, args.version, out_xlsx)
    write_csv(rows_c, out_csv)
    print("Done ✓")


if __name__ == "__main__":
    main()