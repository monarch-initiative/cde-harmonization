#!/usr/bin/env python3
"""
phenx_scraper.py  –  PhenX Toolkit full-hierarchy scraper
==========================================================
Produces:  Collection → Sub-collection → Protocol → Variable (CDE)
           Domain     → Protocol        → Variable (CDE)

Zero hardcoding: collections and domains are discovered by scraping
/collections and /domains at runtime, so new ones appear automatically.

QUICK START
-----------
    pip install requests beautifulsoup4 openpyxl
    python phenx_scraper.py               # all ~1010 protocols
    python phenx_scraper.py --limit 10    # smoke-test
    python phenx_scraper.py --resume      # continue after interrupt

OUTPUT
------
    phenx_hierarchy.xlsx     – README | Protocols | Variables(CDEs) | Reference
    phenx_hierarchy.csv      – flat CDE rows (same data as Variables sheet)
    phenx_checkpoint.json    – resume checkpoint (written next to --out by
                                default, or at --checkpoint-path if given;
                                deleted automatically on a successful run)
"""

import argparse
import csv
import json
import re
import sys
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
BASE            = "https://www.phenxtoolkit.org"
BROWSER_UA      = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

# ── HTTP ──────────────────────────────────────────────────────────────────────

def make_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": BROWSER_UA,
        "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


def get_soup(session, url, delay=0.8, retries=3):
    """Fetch URL with retries; return BeautifulSoup or None."""
    for attempt in range(1, retries + 1):
        time.sleep(delay)
        try:
            r = session.get(url, timeout=30)
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", 60))
                print(f"\n  [rate-limit] sleeping {wait}s", file=sys.stderr)
                time.sleep(wait)
                continue
            r.raise_for_status()
            return BeautifulSoup(r.text, "html.parser")
        except requests.RequestException as e:
            if attempt == retries:
                print(f"\n  [error] {url}: {e}", file=sys.stderr)
                return None
            time.sleep(delay * 3)
    return None


def _abs(href):
    """Make a href absolute."""
    if href.startswith("http"):
        return href
    return BASE + href


def _id_from_url(url, pattern):
    """Extract numeric ID from a URL using a regex pattern string."""
    m = re.search(pattern, url)
    return m.group(1) if m else ""


# ── Phase 0: discover collections and domains from index pages ────────────────

def scrape_collections(session, delay):
    """
    Scrape /collections → list of dicts:
        {id, name, url}
    """
    print("Scraping /collections index …", end=" ", flush=True)
    soup = get_soup(session, f"{BASE}/collections", delay)
    if soup is None:
        raise RuntimeError("Could not fetch /collections")

    results, seen = [], set()
    for a in soup.select("a[href*='/collections/view/']"):
        href = a["href"]
        cid  = _id_from_url(href, r"/collections/view/(\d+)")
        name = a.get_text(strip=True)
        if cid and name and cid not in seen:
            seen.add(cid)
            results.append({"id": cid, "name": name, "url": _abs(href)})

    print(f"{len(results)} collections found")
    return results


def scrape_domains(session, delay):
    """
    Scrape /domains → list of dicts:
        {id, name, url}
    """
    print("Scraping /domains index …", end=" ", flush=True)
    soup = get_soup(session, f"{BASE}/domains", delay)
    if soup is None:
        raise RuntimeError("Could not fetch /domains")

    results, seen = [], set()
    for a in soup.select("a[href*='/domains/view/']"):
        href = a["href"]
        did  = _id_from_url(href, r"/domains/view/(\d+)")
        name = a.get_text(strip=True)
        if did and name and did not in seen:
            seen.add(did)
            results.append({"id": did, "name": name, "url": _abs(href)})

    print(f"{len(results)} domains found")
    return results


# ── Phase 1: discover protocols ───────────────────────────────────────────────

def _extract_proto_links(soup):
    """Return list of (protocol_id, protocol_name) from any page."""
    results, seen = [], set()
    for a in soup.select("a[href*='/protocols/view/']"):
        pid  = _id_from_url(a["href"], r"/protocols/view/(\d+)")
        name = a.get_text(" ", strip=True)
        if pid and name and pid not in seen and not name.startswith("http"):
            seen.add(pid)
            results.append((pid, name))
    return results


def discover_from_domains(session, domains, delay):
    """
    Scrape each domain page for protocol links.
    Returns dict: protocol_id → stub dict
    """
    print("\nPhase 1a: protocols from domain pages …")
    stubs = {}
    for dom in domains:
        print(f"  domain [{dom['id']}] {dom['name'][:45]:<45} … ", end="", flush=True)
        soup = get_soup(session, dom["url"], delay)
        if soup is None:
            print("SKIP")
            continue
        added = 0
        for pid, pname in _extract_proto_links(soup):
            if pid not in stubs:
                stubs[pid] = {
                    "protocol_id":   pid,
                    "protocol_name": pname,
                    "domain_id":     dom["id"],
                    "domain_name":   dom["name"],
                    "domain_url":    dom["url"],
                }
                added += 1
        print(f"+{added:3}  (total {len(stubs)})")
    return stubs


def discover_from_collections(session, collections, delay):
    """
    Scrape each collection page → sub-collection pages → protocol links.

    Returns:
      proto_cols: dict  protocol_id → list of col_entry dicts
      col_stubs:  dict  protocol_id → basic stub
    """
    print("\nPhase 1b: protocols from collection/sub-collection pages …")
    proto_cols = {}   # pid → [col_entry, …]
    col_stubs  = {}   # pid → stub

    for col in collections:
        print(f"  collection [{col['id']}] {col['name'][:40]:<40} … ", end="", flush=True)
        soup = get_soup(session, col["url"], delay)
        if soup is None:
            print("SKIP")
            continue

        # Discover sub-collections listed on this page
        sub_cols, seen_sc = [], set()
        for a in soup.select("a[href*='/sub-collections/view/']"):
            scid = _id_from_url(a["href"], r"/sub-collections/view/(\d+)")
            if scid and scid not in seen_sc:
                seen_sc.add(scid)
                sub_cols.append({
                    "id":   scid,
                    "name": a.get_text(strip=True),
                    "url":  _abs(a["href"]),
                })

        print(f"{len(sub_cols)} sub-cols … ", end="", flush=True)
        new_p = 0

        for sc in sub_cols:
            sc_soup = get_soup(session, sc["url"], delay)
            if sc_soup is None:
                continue
            for pid, pname in _extract_proto_links(sc_soup):
                entry = {
                    "collection_id":       col["id"],
                    "collection_name":     col["name"],
                    "collection_url":      col["url"],
                    "sub_collection_id":   sc["id"],
                    "sub_collection_name": sc["name"],
                    "sub_collection_url":  sc["url"],
                }
                # append (a protocol can be in multiple sub-collections)
                proto_cols.setdefault(pid, []).append(entry)

                if pid not in col_stubs:
                    col_stubs[pid] = {
                        "protocol_id":   pid,
                        "protocol_name": pname,
                        "domain_id":     "",
                        "domain_name":   "",
                        "domain_url":    "",
                    }
                    new_p += 1

        print(f"+{new_p} new protocols")

    return proto_cols, col_stubs


# ── Phase 2: protocol detail + variable parsing ───────────────────────────────

def parse_variables(soup):
    """
    Parse the Variables table on a protocol page.

    HTML pattern (2 rows per variable):
      Row A (group/name):  PX030703_Age_FirstCigarette_Adult_First_Time | "" | "" | ""
      Row B (data):        ""  |  PX030703030000  |  description text  |  dbGaP value

    Correct column mapping:
      col 0 = Variable Name (group identifier, e.g. PX030703_…)
      col 1 = Variable ID   (12-digit, e.g. PX030703030000)
      col 2 = Variable Description
      col 3 = dbGaP Mapping
    """
    # Locate the variable table
    var_tbl = None
    for tag in soup.find_all(["h5", "h6", "strong"]):
        if tag.get_text(strip=True) in ("Variables", "## Variables"):
            var_tbl = tag.find_next("table")
            break
    if var_tbl is None:
        for tbl in soup.find_all("table"):
            hdrs = [th.get_text(strip=True) for th in tbl.find_all("th")]
            if "Variable Name" in hdrs or "Variable ID" in hdrs:
                var_tbl = tbl
                break
    if var_tbl is None:
        return []

    variables     = []
    current_group = ""

    for tr in var_tbl.find_all("tr"):
        tds = tr.find_all(["td", "th"])
        if not tds:
            continue
        raw = [td.get_text(" ", strip=True) for td in tds]

        # Skip column-header row
        if raw[0] in ("Variable Name", "Variable ID", "Variable Description"):
            continue

        col0 = raw[0]
        col1 = raw[1] if len(raw) > 1 else ""
        col2 = raw[2] if len(raw) > 2 else ""
        col3 = raw[3] if len(raw) > 3 else ""

        # Row A: Variable Name (group) — PX + 6-digit protocol code + underscore suffix
        #        col0 matches PX\d{6}_ and col1 is blank
        is_group = (
            bool(col0)
            and re.match(r"^PX\d{6}_", col0)
            and not col1.strip()
        )

        # Row B: Variable data — col0 blank, col1 is 12-digit PX ID
        is_data = (
            not col0.strip()
            and bool(col1)
            and re.match(r"^PX\d{12}$", col1)
        )

        # Edge-case: single-row format (col0=name, col1=id, col2=desc, col3=dbgap)
        is_single = (
            not is_group and not is_data
            and bool(col0)
            and bool(col1)
            and re.match(r"^PX\d{6}_", col0)
            and re.match(r"^PX\d{12}$", col1)
        )

        if is_group:
            current_group = col0

        elif is_data:
            dbgap = re.sub(r"\s*Variable Mapping\s*", "mapped", col3, flags=re.I)
            dbgap = dbgap.replace("N/A", "").strip()
            variables.append({
                "var_name":        current_group,   # e.g. PX030703_Age_FirstCigarette_Adult_First_Time
                "var_id":          col1,            # e.g. PX030703030000
                "var_description": col2,            # question text
                "dbgap_mapping":   dbgap,
            })

        elif is_single:
            dbgap = re.sub(r"\s*Variable Mapping\s*", "mapped", col3, flags=re.I)
            dbgap = dbgap.replace("N/A", "").strip()
            variables.append({
                "var_name":        col0,
                "var_id":          col1,
                "var_description": col2,
                "dbgap_mapping":   dbgap,
            })

    return variables


def fetch_protocol(session, pid, delay, dom_by_id):
    """Fetch a protocol detail page; return detail dict or None."""
    url  = f"{BASE}/protocols/view/{pid}"
    soup = get_soup(session, url, delay)
    if soup is None:
        return None

    d = {"protocol_id": pid, "protocol_url": url}

    # Name
    h1 = soup.find("h1")
    d["protocol_name"] = (h1.get_text(" ", strip=True)
                           .replace("Protocol -", "").strip()) if h1 else ""

    # Instrument / source
    src = soup.find(lambda t: t.name in ("h5","h6","strong")
                              and "Protocol Name from Source" in t.get_text())
    d["instrument"] = (src.find_next_sibling().get_text(strip=True)[:300]
                       if src and src.find_next_sibling() else "")

    # Description
    desc = soup.find(lambda t: t.name in ("h5","h6","strong")
                               and t.get_text(strip=True) == "Description")
    d["description"] = (desc.find_next_sibling().get_text(strip=True)[:600]
                        if desc and desc.find_next_sibling() else "")

    # Mode of administration
    mode = soup.find(lambda t: t.name in ("h5","h6","strong","p")
                               and "Mode of Administration" in t.get_text())
    d["mode"] = (mode.find_next_sibling().get_text(strip=True)[:200]
                 if mode and mode.find_next_sibling() else "")

    # Lifestage
    ls = soup.find(lambda t: t.name in ("h5","h6","strong")
                             and t.get_text(strip=True) == "Lifestage")
    d["lifestage"] = (ls.find_next_sibling().get_text(strip=True)[:100]
                      if ls and ls.find_next_sibling() else "")

    # Availability
    av = soup.find("a", href=re.compile(r"#tabsource"))
    d["availability"] = av.get_text(strip=True)[:120] if av else ""

    # Domain — from page links, resolved against discovered domains
    d["domain_id"] = d["domain_name"] = d["domain_url"] = ""
    for a in soup.select("a[href*='/domains/view/']"):
        did = _id_from_url(a["href"], r"/domains/view/(\d+)$")
        if did and did in dom_by_id:
            d["domain_id"]   = did
            d["domain_name"] = dom_by_id[did]["name"]
            d["domain_url"]  = dom_by_id[did]["url"]
            break

    # LOINC
    d["loinc_code"] = d["loinc_name"] = ""
    for tr in soup.select("table tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all(["td","th"])]
        if len(cells) >= 3 and "LOINC" in cells[0]:
            d["loinc_name"] = cells[1]
            d["loinc_code"] = cells[2]
            break

    d["variables"] = parse_variables(soup)
    return d


# ── Checkpoint ────────────────────────────────────────────────────────────────

def save_ckpt(done, rows_p, rows_c, checkpoint_path):
    with open(checkpoint_path, "w") as f:
        json.dump({"done": list(done), "protocols": rows_p, "cdes": rows_c}, f)

def load_ckpt(checkpoint_path):
    if not Path(checkpoint_path).exists():
        return set(), [], []
    with open(checkpoint_path) as f:
        d = json.load(f)
    return set(d["done"]), d["protocols"], d["cdes"]


# ── Excel builder ─────────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
WHITE    = PatternFill("solid", start_color="FFFFFF")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BDY_FONT = Font(name="Arial", size=9)
BLD_FONT = Font(name="Arial", bold=True, size=9)
LNK_FONT = Font(name="Arial", size=9, color="1155CC", underline="single")
thin     = Side(style="thin", color="BFBFBF")
BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP     = Alignment(vertical="top", wrap_text=True)

def _h(cell, text):
    cell.value = text; cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def _b(cell, val, alt=False, bold=False, link=False):
    cell.value = val
    cell.font  = LNK_FONT if link else (BLD_FONT if bold else BDY_FONT)
    cell.fill  = ALT_FILL if alt else WHITE
    cell.alignment = WRAP; cell.border = BORDER

def _w(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def _title(ws, text, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws.cell(1, 1)
    c.value = text
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def build_excel(rows_p, rows_cde, collections, domains, out_path):
    wb = Workbook()

    # ── README ────────────────────────────────────────────────────────────────
    ws0 = wb.active; ws0.title = "README"
    ws0.column_dimensions["A"].width = 26
    ws0.column_dimensions["B"].width = 90
    meta = [
        ("Source",            BASE),
        ("Scraped",           time.strftime("%Y-%m-%d")),
        ("Collections found", str(len(collections))),
        ("Domains found",     str(len(domains))),
        ("Protocols",         str(len(rows_p))),
        ("CDEs / Variables",  str(len(rows_cde))),
        ("Hierarchy",         "Collection → Sub-collection → Protocol → Variable (CDE)  "
                              "AND  Domain → Protocol → Variable (CDE)"),
        ("Sheets",            "README | Protocols | Variables (CDEs) | Reference"),
        ("License",           "Creative Commons Attribution 4.0 (CC BY 4.0)"),
        ("Attribution",       "The Web-based PhenX Toolkit currently receives funding from NIH; "
                              "official versions at www.phenxtoolkit.org"),
        ("Official downloads",f"{BASE}/resources/download"),
    ]
    for r,(k,v) in enumerate(meta, 1):
        ws0.cell(r,1).value = k; ws0.cell(r,1).font = Font(name="Arial",bold=True,size=10)
        ws0.cell(r,2).value = v; ws0.cell(r,2).font = Font(name="Arial",size=10)

    # ── Protocols ─────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Protocols")
    ws1.freeze_panes = "A3"
    P_HDRS = [
        "Protocol ID","Protocol Name","Instrument / Source",
        "Description","Mode","Lifestage","Availability",
        "Domain ID","Domain Name","Domain URL",
        "Sub-collection ID","Sub-collection Name","Sub-collection URL",
        "Collection ID","Collection Name","Collection URL",
        "LOINC Code","LOINC Name","# Variables","Protocol URL",
    ]
    _title(ws1, f"PhenX Toolkit – Protocols ({len(rows_p)})  |  {BASE}", len(P_HDRS))
    ws1.row_dimensions[2].height = 18
    for c,h in enumerate(P_HDRS,1): _h(ws1.cell(2,c), h)
    for i,w in enumerate([12,52,55,65,30,20,40,10,40,50,12,40,50,10,40,50,12,42,10,50],1):
        _w(ws1, i, w)
    LINK_P = {10,13,16,20}
    for i,p in enumerate(rows_p):
        r=i+3; alt=i%2==1
        vals = [
            p.get("protocol_id",""),   p.get("protocol_name",""),
            p.get("instrument",""),    p.get("description",""),
            p.get("mode",""),          p.get("lifestage",""),
            p.get("availability",""),
            p.get("domain_id",""),     p.get("domain_name",""),    p.get("domain_url",""),
            p.get("sub_collection_id",""),  p.get("sub_collection_name",""), p.get("sub_collection_url",""),
            p.get("collection_id",""),      p.get("collection_name",""),     p.get("collection_url",""),
            p.get("loinc_code",""),    p.get("loinc_name",""),
            str(len(p.get("variables",[]))),
            p.get("protocol_url",""),
        ]
        for c,v in enumerate(vals,1):
            _b(ws1.cell(r,c), v, alt, bold=(c==2), link=(c in LINK_P))
        ws1.row_dimensions[r].height = 32

    # ── Variables / CDEs ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Variables (CDEs)")
    ws2.freeze_panes = "A3"
    V_HDRS = [
        "Variable ID (CDE ID)","Variable Name","Variable Description",
        "dbGaP Mapping",
        "Protocol ID","Protocol Name","Protocol URL",
        "Domain ID","Domain Name","Domain URL",
        "Sub-collection ID","Sub-collection Name","Sub-collection URL",
        "Collection ID","Collection Name","Collection URL",
        "LOINC Code","LOINC Name",
    ]
    _title(ws2, f"PhenX Toolkit – Variables / CDEs ({len(rows_cde)})  |  {BASE}", len(V_HDRS))
    ws2.row_dimensions[2].height = 18
    for c,h in enumerate(V_HDRS,1): _h(ws2.cell(2,c), h)
    for i,w in enumerate([22,46,65,12,12,52,50,10,40,50,12,40,50,10,40,50,12,42],1):
        _w(ws2, i, w)
    LINK_V = {7,10,13,16}
    for i,v in enumerate(rows_cde):
        r=i+3; alt=i%2==1
        vals = [
            v["var_id"],          v["var_name"],
            v["var_description"], v["dbgap_mapping"],
            v["protocol_id"],     v["protocol_name"],   v["protocol_url"],
            v["domain_id"],       v["domain_name"],     v["domain_url"],
            v["sub_collection_id"],   v["sub_collection_name"], v["sub_collection_url"],
            v["collection_id"],       v["collection_name"],     v["collection_url"],
            v["loinc_code"],      v["loinc_name"],
        ]
        for c,val in enumerate(vals,1):
            _b(ws2.cell(r,c), val, alt, bold=(c==1), link=(c in LINK_V))
        ws2.row_dimensions[r].height = 28

    # ── Reference (discovered collections + domains) ──────────────────────────
    ws3 = wb.create_sheet("Reference")
    ws3.freeze_panes = "A2"
    for c,h in enumerate(["Type","ID","Name","URL"],1): _h(ws3.cell(1,c), h)
    ws3.row_dimensions[1].height = 18
    for i,w in enumerate([14,12,62,58],1): _w(ws3,i,w)
    ref = (
        [("Collection", c["id"], c["name"], c["url"]) for c in collections]
        + [("Domain",     d["id"], d["name"], d["url"]) for d in domains]
    )
    for i,(typ,rid,name,url) in enumerate(ref):
        r=i+2; alt=i%2==1
        _b(ws3.cell(r,1), typ,  alt)
        _b(ws3.cell(r,2), rid,  alt)
        _b(ws3.cell(r,3), name, alt, bold=True)
        _b(ws3.cell(r,4), url,  alt, link=True)
        ws3.row_dimensions[r].height = 20

    wb.save(out_path)
    print(f"Saved Excel → {out_path}")


def write_csv(rows_cde, path):
    if not rows_cde:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_cde[0].keys()))
        w.writeheader(); w.writerows(rows_cde)
    print(f"Saved CSV  → {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Scrape PhenX Toolkit → Collection→Domain→Protocol→CDE hierarchy",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--limit",            type=int,   default=0,
                    help="Max protocols to detail-fetch (0 = all)")
    ap.add_argument("--delay",            type=float, default=0.8,
                    help="Seconds between requests (default 0.8)")
    ap.add_argument("--out",              default="phenx_hierarchy.xlsx",
                    help="Output Excel path")
    ap.add_argument("--resume",           action="store_true",
                    help="Resume from checkpoint file")
    ap.add_argument("--checkpoint-every", type=int,   default=50,
                    help="Save checkpoint every N protocols (default 50)")
    ap.add_argument("--checkpoint-path",  default=None,
                    help="Path to checkpoint JSON file "
                         "(default: <out-dir>/phenx_checkpoint.json)")
    args = ap.parse_args()

    out_xlsx  = Path(args.out)
    out_csv   = out_xlsx.with_suffix(".csv")
    ckpt_path = (Path(args.checkpoint_path) if args.checkpoint_path
                 else out_xlsx.parent / "phenx_checkpoint.json")
    ckpt_path.parent.mkdir(parents=True, exist_ok=True)
    session   = make_session()

    # ── Phase 0: discover collections and domains (no hardcoding) ────────────
    print("=" * 60)
    print("Phase 0: discovering collections and domains …")
    collections = scrape_collections(session, args.delay)
    domains     = scrape_domains(session, args.delay)
    dom_by_id   = {d["id"]: d for d in domains}
    print(f"  → {len(collections)} collections, {len(domains)} domains\n")

    # ── Load checkpoint ───────────────────────────────────────────────────────
    done_ids, rows_p, rows_cde = set(), [], []
    if args.resume:
        done_ids, rows_p, rows_cde = load_ckpt(ckpt_path)
        print(f"Resumed: {len(done_ids)} protocols done, {len(rows_cde)} CDEs loaded\n")

    # ── Phase 1: discover all protocol IDs ───────────────────────────────────
    dom_stubs              = discover_from_domains(session, domains, args.delay)
    proto_cols, col_stubs  = discover_from_collections(session, collections, args.delay)

    # Merge: domain stubs are authoritative for domain fields
    all_stubs = dict(col_stubs)
    for pid, stub in dom_stubs.items():
        if pid in all_stubs:
            all_stubs[pid].update({
                "domain_id":   stub["domain_id"],
                "domain_name": stub["domain_name"],
                "domain_url":  stub["domain_url"],
            })
        else:
            all_stubs[pid] = stub

    stubs = [s for s in all_stubs.values() if s["protocol_id"] not in done_ids]
    if args.limit:
        stubs = stubs[:args.limit]

    print(f"\n{'='*60}")
    print(f"Phase 2: fetching details for {len(stubs)} protocols "
          f"(already done: {len(done_ids)}) …\n")

    for i, stub in enumerate(stubs, 1):
        pid = stub["protocol_id"]
        print(f"  [{i:4}/{len(stubs)}] {pid} ", end="", flush=True)

        det = fetch_protocol(session, pid, args.delay, dom_by_id)
        if det is None:
            print("SKIP")
            continue

        # Backfill domain from stub if page didn't resolve it
        if not det.get("domain_id") and stub.get("domain_id"):
            det["domain_id"]   = stub["domain_id"]
            det["domain_name"] = stub["domain_name"]
            det["domain_url"]  = stub["domain_url"]

        # Attach first collection membership to the protocol row
        col_entries = proto_cols.get(pid, [{}])
        best        = col_entries[0]
        det["sub_collection_id"]   = best.get("sub_collection_id","")
        det["sub_collection_name"] = best.get("sub_collection_name","")
        det["sub_collection_url"]  = best.get("sub_collection_url","")
        det["collection_id"]       = best.get("collection_id","")
        det["collection_name"]     = best.get("collection_name","")
        det["collection_url"]      = best.get("collection_url","")

        rows_p.append(det)
        done_ids.add(pid)

        # Emit one CDE row per variable × collection membership
        for col_entry in (col_entries if col_entries else [{}]):
            for var in det.get("variables", []):
                rows_cde.append({
                    "var_id":              var["var_id"],
                    "var_name":            var["var_name"],
                    "var_description":     var["var_description"],
                    "dbgap_mapping":       var["dbgap_mapping"],
                    "protocol_id":         det["protocol_id"],
                    "protocol_name":       det["protocol_name"],
                    "protocol_url":        det["protocol_url"],
                    "domain_id":           det.get("domain_id",""),
                    "domain_name":         det.get("domain_name",""),
                    "domain_url":          det.get("domain_url",""),
                    "sub_collection_id":   col_entry.get("sub_collection_id",""),
                    "sub_collection_name": col_entry.get("sub_collection_name",""),
                    "sub_collection_url":  col_entry.get("sub_collection_url",""),
                    "collection_id":       col_entry.get("collection_id",""),
                    "collection_name":     col_entry.get("collection_name",""),
                    "collection_url":      col_entry.get("collection_url",""),
                    "loinc_code":          det.get("loinc_code",""),
                    "loinc_name":          det.get("loinc_name",""),
                })

        n = len(det.get("variables",[]))
        print(f"→ {det.get('protocol_name','')[:45]}  ({n} vars)")

        if i % args.checkpoint_every == 0:
            save_ckpt(done_ids, rows_p, rows_cde, ckpt_path)
            print(f"  ✓ checkpoint  ({len(rows_p)} protocols, {len(rows_cde)} CDEs)")

    # ── Write outputs ─────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"Total: {len(rows_p)} protocols, {len(rows_cde)} CDEs")
    build_excel(rows_p, rows_cde, collections, domains, out_xlsx)
    write_csv(rows_cde, out_csv)
    if ckpt_path.exists():
        ckpt_path.unlink()
    print("Done ✓")


if __name__ == "__main__":
    main()